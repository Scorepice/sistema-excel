from flask import Flask, render_template, request, send_file, redirect, flash, url_for, abort, jsonify, session
import pandas as pd
import sqlite3
import os
import re
import sys
import html
import datetime
import unicodedata
from fpdf import FPDF
from werkzeug.exceptions import RequestEntityTooLarge


def app_base_dir():
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.abspath(os.path.dirname(__file__))


def resource_path(relative_path):
    base_path = getattr(sys, '_MEIPASS', os.path.abspath(os.path.dirname(__file__)))
    return os.path.join(base_path, relative_path)


BASE_DIR = app_base_dir()

app = Flask(
    __name__,
    template_folder=resource_path('templates'),
    static_folder=resource_path('static'),
)

# Secret key: prefer env var for production; generate a temporary one if missing.
import secrets as _secrets
_env_secret = os.environ.get('SISTEMA_EXCEL_SECRET') or os.environ.get('SISTEMA_EXCEL_SECRET_KEY')
if _env_secret:
    app.secret_key = _env_secret
else:
    # fallback to a generated key but warn the operator; not suitable for multi-instance sessions.
    app.secret_key = _secrets.token_urlsafe(32)
    print('Warning: SISTEMA_EXCEL_SECRET not set — using generated secret key (set env var for stability).')
DB_NAME = os.path.join(BASE_DIR, 'database.db')
DEFAULT_MODULE = 'rdm_abiertas'
MAX_STORED_ROWS_PER_MODULE = int(os.getenv('MAX_STORED_ROWS_PER_MODULE', '100000'))
MAX_TABLE_PAGE_SIZE = int(os.getenv('MAX_TABLE_PAGE_SIZE', '5000'))
MAX_UPLOAD_MB = int(os.getenv('MAX_UPLOAD_MB', '100'))
app.config['MAX_CONTENT_LENGTH'] = MAX_UPLOAD_MB * 1024 * 1024
DATE_MIN_YEAR = int(os.getenv('DATE_MIN_YEAR', '2000'))
DATE_MAX_YEAR = int(os.getenv('DATE_MAX_YEAR', '2100'))

MODULES = {
    'rdm_abiertas': {'nombre': 'RDM ABIERTAS'},
    'rdms': {'nombre': 'RDMs'},
    'estatus_importacion': {'nombre': 'ESTATUS IMPORTACION'},
    'oc_comprometidas': {'nombre': 'OC COMPROMETIDAS'},
    'ordenes_pendientes': {'nombre': 'ORDENES PENDIENTES'},
    'rdms_valoradas': {'nombre': 'RDMS valoradas'},
    'rdms_x_valorar': {'nombre': 'RDMS x VALORAR'},
    'consulta_oc': {'nombre': 'CONSULTA O C'},
}


def get_db_connection():
    conn = sqlite3.connect(DB_NAME)
    conn.row_factory = sqlite3.Row
    return conn


def normalizar_texto(valor):
    texto = unicodedata.normalize('NFKD', str(valor))
    texto = ''.join(caracter for caracter in texto if not unicodedata.combining(caracter))
    return texto.lower().replace('.', '').replace(' ', '').strip()


def table_name_for_module(modulo):
    if modulo == 'rdm_abiertas':
        return 'registros'
    return f"registros_{modulo}"


def get_module_or_404(modulo):
    if modulo not in MODULES:
        abort(404)
    return MODULES[modulo]


def table_exists(conn, table_name):
    cursor = conn.cursor()
    cursor.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name=?",
        (table_name,),
    )
    return cursor.fetchone() is not None


def validate_csrf_token(token):
    return token and session.get('_csrf_token') == token


def detectar_tipo_campo(columna):
    nombre = normalizar_texto(columna)

    if 'estado' in nombre or 'esatad' in nombre:
        return 'estado_select'
    if 'prioridad' in nombre:
        return 'prioridad_select'
    if any(token in nombre for token in ['fecha', 'emision', 'desde', 'hasta']):
        return 'date'
    if any(token in nombre for token in ['cant', 'monto', 'valor', 'precio', 'total', 'desde', 'hasta']):
        return 'number'
    return 'text'


def construir_campos_formulario(columnas):
    return [{'name': col, 'kind': detectar_tipo_campo(col)} for col in columnas]


def should_parse_as_date(column_name):
    nombre = normalizar_texto(column_name)
    return any(token in nombre for token in ['fecha', 'emision', 'desde', 'hasta'])


def get_dashboard_date_column(modulo, df):
    if modulo == 'rdm_abiertas':
        for candidate in ['# parte', '#parte', 'descripci�n', 'descripcion']:
            if candidate in df.columns:
                return candidate

    candidates = [c for c in df.columns if should_parse_as_date(c)]
    if candidates:
        return candidates[0]

    best_column = None
    best_ratio = 0.0
    for column in df.columns:
        parsed_ratio = pd.to_datetime(df[column].apply(parse_date_like_value), errors='coerce').notna().mean()
        if parsed_ratio > best_ratio:
            best_ratio = parsed_ratio
            best_column = column

    return best_column


def parse_date_like_value(value):
    if value is None:
        return None

    # Evita errores con NaN/NaT provenientes de pandas/openpyxl.
    try:
        if pd.isna(value):
            return None
    except TypeError:
        pass

    if isinstance(value, pd.Timestamp):
        if not (DATE_MIN_YEAR <= value.year <= DATE_MAX_YEAR):
            return None
        return value.strftime('%Y-%m-%d')

    if isinstance(value, (datetime.datetime, datetime.date)):
        fecha = pd.Timestamp(value)
        if not (DATE_MIN_YEAR <= fecha.year <= DATE_MAX_YEAR):
            return None
        return fecha.strftime('%Y-%m-%d')

    if isinstance(value, (int, float)) and not isinstance(value, bool):
        # 0 y negativos en campos fecha suelen representar vacio/invalido.
        if value <= 0:
            return None

        # Serial de fecha de Excel (dias desde 1899-12-30).
        if value <= 60000:
            fecha = pd.to_datetime(value, unit='D', origin='1899-12-30', errors='coerce')
            if pd.notna(fecha) and DATE_MIN_YEAR <= fecha.year <= DATE_MAX_YEAR:
                return fecha.strftime('%Y-%m-%d')
        return None

    if isinstance(value, str):
        texto = value.strip()
        if not texto:
            return None

        if texto in {'0', '0.0', '00/00/0000', '0000-00-00'}:
            return None

        if texto.isdigit():
            numero = int(texto)
            if numero <= 0:
                return None

            # Trata numeros puros como serial de Excel solo si caen en un rango razonable.
            if numero <= 60000:
                fecha = pd.to_datetime(numero, unit='D', origin='1899-12-30', errors='coerce')
                if pd.notna(fecha) and 1900 <= fecha.year <= 2100:
                    return fecha.strftime('%Y-%m-%d')
            return None

        if re.fullmatch(r'\d{4}-\d{2}-\d{2}', texto):
            fecha = pd.to_datetime(texto, format='%Y-%m-%d', errors='coerce')
            if pd.notna(fecha) and DATE_MIN_YEAR <= fecha.year <= DATE_MAX_YEAR:
                return fecha.strftime('%Y-%m-%d')

        if any(separador in texto for separador in ['/', '-', ':']) or any(mes in texto.lower() for mes in ['ene', 'feb', 'mar', 'abr', 'may', 'jun', 'jul', 'ago', 'sep', 'oct', 'nov', 'dic']):
            fecha = pd.to_datetime(texto, errors='coerce', dayfirst=True)
            if pd.notna(fecha) and DATE_MIN_YEAR <= fecha.year <= DATE_MAX_YEAR:
                return fecha.strftime('%Y-%m-%d')

    return None


def align_dataframe_to_existing_table(df, conn, table_name):
    cursor = conn.cursor()
    cursor.execute(f'PRAGMA table_info("{table_name}")')
    existing_columns = [col[1] for col in cursor.fetchall()]

    if not existing_columns:
        return df

    normalized_existing = {normalizar_texto(col): col for col in existing_columns}
    rename_map = {}
    for col in df.columns:
        normalized_col = normalizar_texto(col)
        if normalized_col in normalized_existing:
            rename_map[col] = normalized_existing[normalized_col]

    if rename_map:
        df = df.rename(columns=rename_map)

    matched_columns = [col for col in existing_columns if col in df.columns]
    if not matched_columns:
        raise ValueError('El Excel no coincide con la estructura esperada del modulo.')

    # Evita cargar archivos de estructura distinta que causen corrimiento de datos.
    min_base = max(1, min(len(existing_columns), len(df.columns)))
    compatibility_ratio = len(matched_columns) / min_base
    if compatibility_ratio < 0.60:
        raise ValueError(
                'El Excel parece pertenecer a otra estructura/modulo. '
                f'Compatibilidad detectada: {compatibility_ratio:.0%}. '
                'Revisa que corresponda al modulo seleccionado.'
            )

    # Si el Excel trae columnas nuevas, agregarlas a la tabla para no perder informacion.
    new_columns = [col for col in df.columns if col not in existing_columns]
    for col in new_columns:
        safe_col = str(col).replace('"', '""')
        conn.execute(f'ALTER TABLE "{table_name}" ADD COLUMN "{safe_col}" TEXT')
        existing_columns.append(col)

    for col in existing_columns:
        if col not in df.columns:
            df[col] = None

    return df[existing_columns]


def format_table_value(value):
    if value is None:
        return ''

    text = str(value).replace(' 00:00:00', '')
    if text in ['None', 'nan']:
        return ''
    if text.endswith('.0'):
        return text[:-2]
    return text


def worksheet_to_dataframe(worksheet):
    rows = worksheet.iter_rows(values_only=True)

    try:
        encabezados = next(rows)
    except StopIteration:
        return pd.DataFrame()

    encabezados_limpios = []
    for index, encabezado in enumerate(encabezados, start=1):
        texto = ' '.join(str(encabezado).split()) if encabezado is not None else ''
        encabezados_limpios.append(texto if texto else f'Columna_{index}')

    registros = []
    for row in rows:
        if row is None:
            continue

        valores = list(row)
        if len(valores) < len(encabezados_limpios):
            valores.extend([None] * (len(encabezados_limpios) - len(valores)))
        elif len(valores) > len(encabezados_limpios):
            extras = len(valores) - len(encabezados_limpios)
            encabezados_limpios.extend([f'Columna_{len(encabezados_limpios) + i + 1}' for i in range(extras)])

        if all(valor is None or str(valor).strip() == '' for valor in valores):
            continue

        fila_dict = dict(zip(encabezados_limpios, valores))

        # Si una fila repite el encabezado, saltarla.
        if all(normalizar_texto(fila_dict.get(col, '')) == normalizar_texto(col) for col in encabezados_limpios):
            continue

        registros.append(fila_dict)

    if not registros:
        return pd.DataFrame()

    return pd.DataFrame(registros)


def enforce_table_row_limit(conn, table_name, max_rows=MAX_STORED_ROWS_PER_MODULE):
    total_rows = conn.execute(f'SELECT COUNT(*) FROM "{table_name}"').fetchone()[0]
    overflow = total_rows - max_rows
    if overflow <= 0:
        return 0, total_rows

    # Borra los registros más antiguos (rowid menor) para conservar los más recientes.
    conn.execute(
        f'''DELETE FROM "{table_name}"
            WHERE rowid IN (
                SELECT rowid FROM "{table_name}" ORDER BY rowid ASC LIMIT ?
            )''',
        (overflow,),
    )
    conn.commit()

    remaining_rows = conn.execute(f'SELECT COUNT(*) FROM "{table_name}"').fetchone()[0]
    return overflow, remaining_rows


def leer_excel_completo(archivo):
    from openpyxl import load_workbook

    workbook = load_workbook(archivo, data_only=True, read_only=True)
    frames = []

    for worksheet in workbook.worksheets:
        df_hoja = worksheet_to_dataframe(worksheet)
        if not df_hoja.empty:
            frames.append(df_hoja)

    workbook.close()

    if not frames:
        return pd.DataFrame()

    return pd.concat(frames, ignore_index=True, sort=False)


@app.context_processor
def inject_globals():
    # expose modules and a csrf token generator to templates
    def _csrf_token():
        if '_csrf_token' not in session:
            session['_csrf_token'] = _secrets.token_urlsafe(16)
        return session['_csrf_token']

    return {'modulos': MODULES, 'csrf_token': _csrf_token}


@app.route('/')
def selector_modulo():
    return render_template('module_selector.html', modules=MODULES)


@app.route('/modulo/<modulo>')
def index(modulo):
    modulo_info = get_module_or_404(modulo)
    return render_template('index.html', modulo=modulo, module_name=modulo_info['nombre'])


@app.route('/modulo/<modulo>/subir', methods=['POST'])
def subir_excel(modulo):
    modulo_info = get_module_or_404(modulo)
    archivo = request.files.get('archivo')

    if not archivo:
        flash('Selecciona un archivo para procesar.', 'warning')
        return redirect(url_for('index', modulo=modulo))

    try:
        df = leer_excel_completo(archivo)

        if df.empty:
            flash('El archivo no contiene filas validas para procesar.', 'warning')
            return redirect(url_for('index', modulo=modulo))

        for col in df.columns:
            if should_parse_as_date(col):
                df[col] = df[col].apply(parse_date_like_value)

        conn = sqlite3.connect(DB_NAME)
        table_name = table_name_for_module(modulo)
        if table_exists(conn, table_name):
            df = align_dataframe_to_existing_table(df, conn, table_name)

        df.to_sql(table_name, conn, if_exists='append', index=False)
        eliminados, total_actual = enforce_table_row_limit(conn, table_name, MAX_STORED_ROWS_PER_MODULE)
        conn.close()

        if eliminados > 0:
            flash(
                (
                    f"Excel del modulo '{modulo_info['nombre']}' guardado con exito. "
                    f"Se eliminaron {eliminados} registros antiguos para mantener el limite de "
                    f"{MAX_STORED_ROWS_PER_MODULE} (total actual: {total_actual})."
                ),
                'warning',
            )
        else:
            flash(
                (
                    f"Excel del modulo '{modulo_info['nombre']}' guardado con exito leyendo todas las hojas. "
                    f"Total actual: {total_actual}."
                ),
                'success',
            )
        return redirect(url_for('ver_datos', modulo=modulo))
    except Exception as e:
        flash(f'Error al procesar el archivo: {str(e)}', 'danger')
        return redirect(url_for('index', modulo=modulo))


@app.errorhandler(RequestEntityTooLarge)
def handle_large_upload(_error):
    flash(f'El archivo excede el tamano maximo permitido ({MAX_UPLOAD_MB} MB).', 'danger')
    return redirect(request.referrer or url_for('selector_modulo'))


@app.route('/modulo/<modulo>/datos')
def ver_datos(modulo):
    modulo_info = get_module_or_404(modulo)
    table_name = table_name_for_module(modulo)

    conn = get_db_connection()
    try:
        if not table_exists(conn, table_name):
            flash(f"El modulo '{modulo_info['nombre']}' aun no tiene datos. Sube un Excel primero.", 'warning')
            return render_template(
                'datos.html',
                modulo=modulo,
                module_name=modulo_info['nombre'],
                filas=[],
                columnas=[],
                table_columns=[],
            )

        filas = conn.execute(
            f'SELECT rowid, * FROM "{table_name}" ORDER BY rowid DESC LIMIT {MAX_TABLE_PAGE_SIZE}'
        ).fetchall()
        cursor = conn.cursor()
        cursor.execute(f'PRAGMA table_info("{table_name}")')
        columnas = [col[1] for col in cursor.fetchall()]

        return render_template(
            'datos.html',
            modulo=modulo,
            module_name=modulo_info['nombre'],
            filas=filas,
            columnas=columnas,
            table_columns=[
                *[{'data': i} for i in range(len(columnas))],
                {'data': len(columnas), 'orderable': False, 'searchable': False},
                {
                    'data': len(columnas) + 1,
                    'orderable': False,
                    'searchable': False,
                    'className': 'text-center',
                },
            ],
        )
    finally:
        conn.close()


@app.route('/modulo/<modulo>/datos_json')
def ver_datos_json(modulo):
    get_module_or_404(modulo)
    table_name = table_name_for_module(modulo)

    draw = int(request.args.get('draw', 1))
    start = max(int(request.args.get('start', 0)), 0)
    length = int(request.args.get('length', 100))
    if length <= 0:
        length = 100
    if length > MAX_TABLE_PAGE_SIZE:
        length = MAX_TABLE_PAGE_SIZE

    conn = get_db_connection()
    try:
        if not table_exists(conn, table_name):
            return jsonify({'draw': draw, 'recordsTotal': 0, 'recordsFiltered': 0, 'data': []})

        cursor = conn.cursor()
        cursor.execute(f'PRAGMA table_info("{table_name}")')
        columnas = [col[1] for col in cursor.fetchall()]

        if not columnas:
            return jsonify({'draw': draw, 'recordsTotal': 0, 'recordsFiltered': 0, 'data': []})

        total = conn.execute(f'SELECT COUNT(*) FROM "{table_name}"').fetchone()[0]

        busqueda = request.args.get('search[value]', '').strip()
        where_clause = ''
        where_params = []
        if busqueda:
            like = f'%{busqueda}%'
            filters = [f'CAST("{col}" AS TEXT) LIKE ?' for col in columnas]
            where_clause = ' WHERE ' + ' OR '.join(filters)
            where_params = [like] * len(columnas)

        filtered = total
        if where_clause:
            filtered = conn.execute(
                f'SELECT COUNT(*) FROM "{table_name}"{where_clause}',
                where_params,
            ).fetchone()[0]

        order_column_index = request.args.get('order[0][column]')
        order_dir = request.args.get('order[0][dir]', 'desc').lower()
        order_dir_sql = 'DESC' if order_dir == 'desc' else 'ASC'
        order_clause = ' ORDER BY rowid DESC'

        if order_column_index is not None:
            try:
                idx = int(order_column_index)
                if 0 <= idx < len(columnas):
                    order_clause = f' ORDER BY "{columnas[idx]}" {order_dir_sql}'
            except ValueError:
                pass

        query = (
            f'SELECT rowid, * FROM "{table_name}"'
            f'{where_clause}'
            f'{order_clause}'
            ' LIMIT ? OFFSET ?'
        )
        rows = conn.execute(query, where_params + [length, start]).fetchall()

        data = []
        for row in rows:
            rendered_row = [html.escape(format_table_value(row[col])) for col in columnas]
            edit_url = url_for('editar', modulo=modulo, id=row['rowid'])
            delete_url = url_for('eliminar', modulo=modulo, id=row['rowid'])
            acciones = (
                f'<a href="{edit_url}" class="btn btn-sm btn-warning fw-bold">Editar</a> '
                f'<a href="{delete_url}" class="btn btn-sm btn-danger fw-bold" '
                f'onclick="return confirm(\'¿Seguro que deseas borrar este registro?\')">Borrar</a>'
            )
            rendered_row.append(acciones)
            rendered_row.append(
                f'<input type="checkbox" class="form-check-input row-selector" '
                f'data-id="{row["rowid"]}" value="{row["rowid"]}" '
                'aria-label="Seleccionar registro">'
            )
            data.append(rendered_row)

        return jsonify(
            {
                'draw': draw,
                'recordsTotal': total,
                'recordsFiltered': filtered,
                'data': data,
            }
        )
    finally:
        conn.close()


@app.route('/modulo/<modulo>/dashboard')
def dashboard(modulo):
    modulo_info = get_module_or_404(modulo)
    table_name = table_name_for_module(modulo)

    data_vacia = {
        'labels': [],
        'datasets': [],
        'total_labels': [],
        'total_valores': [],
        'total_colores': [],
    }

    inicio = request.args.get('inicio', '')
    fin = request.args.get('fin', '')

    conn = get_db_connection()
    try:
        if not table_exists(conn, table_name):
            flash(f"El modulo '{modulo_info['nombre']}' aun no tiene datos para graficas.", 'warning')
            return render_template(
                'dashboard.html',
                modulo=modulo,
                module_name=modulo_info['nombre'],
                data_estatus=data_vacia,
                data_depto={'labels': [], 'valores': [], 'colores': []},
                data_cruce={'labels': [], 'datasets': []},
                inicio=inicio,
                fin=fin,
            )

        df = pd.read_sql_query(f'SELECT * FROM "{table_name}"', conn)
    finally:
        conn.close()

    if df.empty:
        return render_template(
            'dashboard.html',
            modulo=modulo,
            module_name=modulo_info['nombre'],
            data_estatus=data_vacia,
            data_depto={'labels': [], 'valores': [], 'colores': []},
            data_cruce={'labels': [], 'datasets': []},
            inicio=inicio,
            fin=fin,
        )

    col_estado = next((c for c in df.columns if 'estad' in c.lower() or 'esatad' in c.lower()), None)
    col_fecha = get_dashboard_date_column(modulo, df)
    col_depto = next((c for c in df.columns if 'departamento' in c.lower() or 'depto' in c.lower()), None)

    if not col_estado or not col_fecha or not col_depto:
        flash('No se detectaron columnas necesarias para graficas (Estado, Fecha y Departamento).', 'danger')
        return render_template(
            'dashboard.html',
            modulo=modulo,
            module_name=modulo_info['nombre'],
            data_estatus=data_vacia,
            data_depto={'labels': [], 'valores': [], 'colores': []},
            data_cruce={'labels': [], 'datasets': []},
            inicio=inicio,
            fin=fin,
        )

    df['Fecha_Real'] = pd.to_datetime(df[col_fecha].apply(parse_date_like_value), errors='coerce')

    if inicio:
        df = df[df['Fecha_Real'] >= pd.to_datetime(inicio)]
    if fin:
        df = df[df['Fecha_Real'] <= pd.to_datetime(fin)]

    if df.empty:
        return render_template(
            'dashboard.html',
            modulo=modulo,
            module_name=modulo_info['nombre'],
            data_estatus=data_vacia,
            data_depto={'labels': [], 'valores': [], 'colores': []},
            data_cruce={'labels': [], 'datasets': []},
            inicio=inicio,
            fin=fin,
        )

    df['Mes'] = df['Fecha_Real'].dt.strftime('%Y-%m').fillna('Sin Fecha')
    df[col_depto] = df[col_depto].fillna('SIN DEPTO').astype(str).str.strip().str.upper()
    df[col_estado] = df[col_estado].fillna('PENDIENTE').astype(str).str.strip().str.upper()

    estado_counts = df[col_estado].value_counts()
    paleta_estados = ['#17659d', '#fd7e14', '#6f42c1', '#20c997', '#e83e8c', '#dc3545', '#0dcaf0', '#ffc107', '#28a745', '#6610f2']
    colores_estado_totales = [paleta_estados[i % len(paleta_estados)] for i in range(len(estado_counts))]
    mapa_colores_estado = {
        str(estado): colores_estado_totales[i]
        for i, estado in enumerate(estado_counts.index.astype(str).tolist())
    }

    pivot_estado = pd.crosstab(df['Mes'], df[col_estado])
    data_estatus = {
        'labels': pivot_estado.index.tolist(),
        'datasets': [],
        'total_labels': estado_counts.index.astype(str).tolist(),
        'total_valores': estado_counts.values.tolist(),
        'total_colores': colores_estado_totales,
    }
    for estado in pivot_estado.columns.astype(str).tolist():
        color_asignado = mapa_colores_estado.get(estado, '#cccccc')
        data_estatus['datasets'].append(
            {
                'label': estado,
                'data': pivot_estado[estado].tolist(),
                'backgroundColor': color_asignado,
                'borderColor': color_asignado,
                'borderWidth': 1,
            }
        )

    deptos_unicos = df[col_depto].unique().tolist()
    paleta = ['#17659d', '#fd7e14', '#6f42c1', '#20c997', '#e83e8c', '#dc3545', '#0dcaf0', '#ffc107', '#28a745', '#6610f2']
    mapa_colores = {depto: paleta[i % len(paleta)] for i, depto in enumerate(deptos_unicos)}

    depto_counts = df[col_depto].value_counts()
    colores_depto = [mapa_colores.get(d, '#cccccc') for d in depto_counts.index]
    data_depto = {
        'labels': depto_counts.index.astype(str).tolist(),
        'valores': depto_counts.values.tolist(),
        'colores': colores_depto,
    }

    pivot = pd.crosstab(df['Mes'], df[col_depto])
    data_cruce = {'labels': pivot.index.tolist(), 'datasets': []}
    for depto in pivot.columns:
        color_asignado = mapa_colores.get(depto, '#cccccc')
        data_cruce['datasets'].append(
            {
                'label': str(depto),
                'data': pivot[depto].tolist(),
                'backgroundColor': color_asignado,
                'borderColor': color_asignado,
                'borderWidth': 1,
            }
        )

    return render_template(
        'dashboard.html',
        modulo=modulo,
        module_name=modulo_info['nombre'],
        data_estatus=data_estatus,
        data_depto=data_depto,
        data_cruce=data_cruce,
        inicio=inicio,
        fin=fin,
    )


@app.route('/modulo/<modulo>/editar/<int:id>', methods=['GET', 'POST'])
def editar(modulo, id):
    modulo_info = get_module_or_404(modulo)
    table_name = table_name_for_module(modulo)

    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    if not table_exists(conn, table_name):
        conn.close()
        flash(f"El modulo '{modulo_info['nombre']}' aun no tiene tabla de datos.", 'warning')
        return redirect(url_for('index', modulo=modulo))

    cursor.execute(f'PRAGMA table_info("{table_name}")')
    columnas_db = [col[1] for col in cursor.fetchall()]

    if request.method == 'POST':
        datos_html = dict(request.form)
        datos_a_actualizar = {}

        for col_db in columnas_db:
            col_db_limpia = normalizar_texto(col_db)
            for key_html, valor in datos_html.items():
                key_html_limpia = normalizar_texto(key_html)
                if col_db_limpia == key_html_limpia:
                    datos_a_actualizar[col_db] = None if valor.strip() == '' else valor
                    break

        if datos_a_actualizar:
            set_clause = ', '.join([f'"{k}" = ?' for k in datos_a_actualizar.keys()])
            valores = list(datos_a_actualizar.values())
            valores.append(id)
            try:
                cursor.execute(f'UPDATE "{table_name}" SET {set_clause} WHERE rowid = ?', valores)
                conn.commit()
                flash('Registro actualizado correctamente.', 'success')
            except Exception as e:
                flash(f'Error al actualizar: {str(e)}', 'danger')
        else:
            flash('No se encontraron datos validos para actualizar.', 'warning')

        conn.close()
        return redirect(url_for('ver_datos', modulo=modulo))

    cursor.execute(f'SELECT rowid, * FROM "{table_name}" WHERE rowid = ?', (id,))
    registro = cursor.fetchone()
    conn.close()

    if registro is None:
        flash('El registro no existe.', 'warning')
        return redirect(url_for('ver_datos', modulo=modulo))

    campos = construir_campos_formulario(columnas_db)
    return render_template(
        'editar.html',
        modulo=modulo,
        module_name=modulo_info['nombre'],
        registro=registro,
        campos=campos,
    )


@app.route('/modulo/<modulo>/agregar', methods=['GET', 'POST'])
def agregar(modulo):
    modulo_info = get_module_or_404(modulo)
    table_name = table_name_for_module(modulo)

    conn = get_db_connection()
    cursor = conn.cursor()

    if not table_exists(conn, table_name):
        conn.close()
        flash(
            f"El modulo '{modulo_info['nombre']}' no tiene estructura aun. Sube un Excel primero.",
            'warning',
        )
        return redirect(url_for('index', modulo=modulo))

    cursor.execute(f'PRAGMA table_info("{table_name}")')
    columnas_info = cursor.fetchall()
    columnas_db = [col[1] for col in columnas_info]

    if request.method == 'POST':
        datos_html = dict(request.form)
        datos_a_guardar = {}

        for col_db in columnas_db:
            col_db_limpia = normalizar_texto(col_db)
            for key_html, valor in datos_html.items():
                key_html_limpia = normalizar_texto(key_html)
                if col_db_limpia == key_html_limpia:
                    datos_a_guardar[col_db] = None if valor.strip() == '' else valor
                    break

        if not datos_a_guardar:
            conn.close()
            flash('No se detectaron campos validos para guardar.', 'warning')
            return redirect(url_for('agregar', modulo=modulo))

        columnas_str = ', '.join([f'"{k}"' for k in datos_a_guardar.keys()])
        placeholders = ', '.join(['?' for _ in datos_a_guardar])
        valores = list(datos_a_guardar.values())

        try:
            cursor.execute(
                f'INSERT INTO "{table_name}" ({columnas_str}) VALUES ({placeholders})',
                valores,
            )
            conn.commit()
            flash('Registro guardado correctamente.', 'success')
        except Exception as e:
            flash(f'Error al guardar en la base de datos: {str(e)}', 'danger')
        finally:
            conn.close()

        return redirect(url_for('ver_datos', modulo=modulo))

    conn.close()
    campos = construir_campos_formulario(columnas_db)
    return render_template(
        'agregar.html',
        modulo=modulo,
        module_name=modulo_info['nombre'],
        campos=campos,
    )


@app.route('/modulo/<modulo>/eliminar/<int:id>', methods=['POST'])
def eliminar(modulo, id):
    get_module_or_404(modulo)
    table_name = table_name_for_module(modulo)

    token = request.form.get('_csrf_token')
    if not validate_csrf_token(token):
        flash('Token CSRF invalido. Operacion cancelada.', 'danger')
        return redirect(url_for('ver_datos', modulo=modulo))

    conn = get_db_connection()
    try:
        if not table_exists(conn, table_name):
            flash('No hay tabla para eliminar registros en este modulo.', 'warning')
            return redirect(url_for('index', modulo=modulo))

        conn.execute(f'DELETE FROM "{table_name}" WHERE rowid = ?', (id,))
        conn.commit()
        flash('Registro eliminado.', 'danger')
    finally:
        conn.close()

    return redirect(url_for('ver_datos', modulo=modulo))


@app.route('/modulo/<modulo>/eliminar_multiples', methods=['POST'])
def eliminar_multiples(modulo):
    get_module_or_404(modulo)
    table_name = table_name_for_module(modulo)

    ids_raw = request.form.getlist('ids')
    token = request.form.get('_csrf_token')
    if not validate_csrf_token(token):
        flash('Token CSRF invalido. Operacion cancelada.', 'danger')
        return redirect(url_for('ver_datos', modulo=modulo))
    if not ids_raw:
        flash('Selecciona al menos un registro para eliminar.', 'warning')
        return redirect(url_for('ver_datos', modulo=modulo))

    ids = []
    for value in ids_raw:
        try:
            ids.append(int(value))
        except (TypeError, ValueError):
            continue

    ids = sorted(set(ids))
    if not ids:
        flash('No se recibieron IDs validos para eliminar.', 'warning')
        return redirect(url_for('ver_datos', modulo=modulo))

    conn = get_db_connection()
    try:
        if not table_exists(conn, table_name):
            flash('No hay tabla para eliminar registros en este modulo.', 'warning')
            return redirect(url_for('index', modulo=modulo))

        placeholders = ', '.join(['?' for _ in ids])
        cursor = conn.execute(
            f'DELETE FROM "{table_name}" WHERE rowid IN ({placeholders})',
            ids,
        )
        conn.commit()

        if cursor.rowcount > 0:
            flash(f'Se eliminaron {cursor.rowcount} registros.', 'danger')
        else:
            flash('No se eliminaron registros (puede que ya no existieran).', 'warning')
    finally:
        conn.close()

    return redirect(url_for('ver_datos', modulo=modulo))


@app.route('/modulo/<modulo>/limpiar_base')
@app.route('/modulo/<modulo>/limpiar_base', methods=['POST'])
def limpiar_base(modulo):
    modulo_info = get_module_or_404(modulo)
    table_name = table_name_for_module(modulo)

    token = request.form.get('_csrf_token')
    if not validate_csrf_token(token):
        flash('Token CSRF invalido. Operacion cancelada.', 'danger')
        return redirect(url_for('index', modulo=modulo))

    conn = get_db_connection()
    try:
        conn.execute(f'DROP TABLE IF EXISTS "{table_name}"')
        conn.commit()
        flash(f"Base del modulo '{modulo_info['nombre']}' limpiada.", 'success')
    except Exception:
        flash('No se pudo limpiar la base de este modulo.', 'danger')
    finally:
        conn.close()

    return redirect(url_for('index', modulo=modulo))


@app.route('/modulo/<modulo>/reporte/<tipo>')
def generar_reporte(modulo, tipo):
    modulo_info = get_module_or_404(modulo)
    table_name = table_name_for_module(modulo)

    conn = get_db_connection()
    try:
        if not table_exists(conn, table_name):
            flash(f"El modulo '{modulo_info['nombre']}' aun no tiene datos para exportar.", 'warning')
            return redirect(url_for('ver_datos', modulo=modulo))

        df = pd.read_sql_query(f'SELECT * FROM "{table_name}"', conn)
    finally:
        conn.close()

    busqueda = request.args.get('q', '').strip()
    if busqueda:
        mask = df.astype(str).apply(lambda x: x.str.contains(busqueda, case=False, na=False)).any(axis=1)
        df = df[mask]

    if df.empty:
        flash('No hay datos que coincidan con la busqueda para exportar.', 'warning')
        return redirect(url_for('ver_datos', modulo=modulo))

    if tipo == 'excel':
        ruta = os.path.join(BASE_DIR, f"reporte_{modulo}.xlsx")
        df.to_excel(ruta, index=False)
        return send_file(ruta, as_attachment=True)

    if tipo == 'pdf':
        pdf = FPDF(orientation='L', format='A3')
        pdf.add_page()

        ruta_logo = os.path.join('static', 'maritime_foot.png')
        if os.path.exists(ruta_logo):
            pdf.image(ruta_logo, x=15, y=3, w=40)

        pdf.set_font('Arial', style='B', size=16)
        titulo = (
            f"Reporte {modulo_info['nombre']}: '{busqueda}'"
            if busqueda
            else f"Reporte General {modulo_info['nombre']}"
        )
        pdf.cell(0, 15, txt=titulo, ln=True, align='C')
        pdf.ln(5)

        anchos = []
        for col in df.columns:
            nombre = re.sub(r'[\s\.]', '', str(col).lower())
            if nombre == 'um':
                anchos.append(10)
            elif nombre == 'rg':
                anchos.append(55)
            elif nombre == 'usuario':
                anchos.append(12)
            elif 'cant' in nombre:
                anchos.append(25)
            elif 'rdm' in nombre or 'prioridad' in nombre:
                anchos.append(16)
            elif nombre in ['desde', 'hasta']:
                anchos.append(14)
            elif nombre in ['fecha', 'parte', 'descripción', 'descripcion']:
                anchos.append(18)
            elif 'estado' in nombre or 'esatado' in nombre:
                anchos.append(22)
            elif 'codigosistemas' in nombre:
                anchos.append(22)
            elif 'departamento' in nombre:
                anchos.append(26)
            elif 'código' in nombre or 'codigo' in nombre:
                anchos.append(68)
            else:
                anchos.append(20)

        pdf.set_font('Arial', style='B', size=8)
        for i, col in enumerate(df.columns):
            texto_col = (
                str(col)
                .replace('“', '"')
                .replace('”', '"')
                .replace('‘', "'")
                .replace('’', "'")
                .replace('–', '-')
                .replace('—', '-')
            )
            texto_col = texto_col.encode('latin-1', errors='ignore').decode('latin-1')
            pdf.cell(anchos[i], 8, texto_col[: int(anchos[i] * 0.75)], border=1, align='C')
        pdf.ln()

        pdf.set_font('Arial', size=7)
        for _, row in df.iterrows():
            for i, val in enumerate(row):
                texto = str(val).replace(' 00:00:00', '')
                if texto in ['nan', 'None']:
                    texto = ''
                if texto.endswith('.0'):
                    texto = texto[:-2]

                texto_limpio = (
                    texto.replace('“', '"')
                    .replace('”', '"')
                    .replace('‘', "'")
                    .replace('’', "'")
                    .replace('–', '-')
                    .replace('—', '-')
                )
                texto_limpio = texto_limpio.encode('latin-1', errors='ignore').decode('latin-1')
                limite = int(anchos[i] * 0.65)
                pdf.cell(anchos[i], 7, texto_limpio[:limite], border=1)
            pdf.ln()

        ruta_pdf = os.path.join(BASE_DIR, f"reporte_{modulo}.pdf")
        pdf.output(ruta_pdf)
        return send_file(ruta_pdf, as_attachment=True)

    flash('Tipo de reporte no valido. Usa excel o pdf.', 'warning')
    return redirect(url_for('ver_datos', modulo=modulo))


# Redirecciones de compatibilidad para enlaces antiguos
@app.route('/datos')
def legacy_datos():
    return redirect(url_for('ver_datos', modulo=DEFAULT_MODULE))


@app.route('/dashboard')
def legacy_dashboard():
    return redirect(url_for('dashboard', modulo=DEFAULT_MODULE))


@app.route('/agregar')
def legacy_agregar():
    return redirect(url_for('agregar', modulo=DEFAULT_MODULE))


@app.route('/editar/<int:id>', methods=['GET', 'POST'])
def legacy_editar(id):
    return redirect(url_for('editar', modulo=DEFAULT_MODULE, id=id))


@app.route('/eliminar/<int:id>')
def legacy_eliminar(id):
    # eliminar now requires POST with CSRF; redirect to datos view instead.
    return redirect(url_for('ver_datos', modulo=DEFAULT_MODULE))


@app.route('/limpiar_base')
def legacy_limpiar_base():
    # limpiar_base now requires POST; redirect to index instead to avoid accidental deletes.
    return redirect(url_for('index', modulo=DEFAULT_MODULE))


@app.route('/reporte/<tipo>')
def legacy_reporte(tipo):
    return redirect(url_for('generar_reporte', modulo=DEFAULT_MODULE, tipo=tipo, q=request.args.get('q', '')))


@app.route('/subir', methods=['POST'])
def legacy_subir():
    return redirect(url_for('index', modulo=DEFAULT_MODULE))


if __name__ == '__main__':
    debug_mode = os.environ.get('SISTEMA_EXCEL_DEBUG', '0') == '1'
    port = int(os.environ.get('SISTEMA_EXCEL_PORT', '5000'))
    app.run(host='127.0.0.1', port=port, debug=debug_mode)
